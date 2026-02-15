# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
import shutil

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "税理士法人杉井総合会計事務所" / "提案" / "【ドワンゴ様】お見積り.xlsx"
OUT = ROOT / "税理士法人杉井総合会計事務所" / "提案" / "伴走支援プラン見積もり（3パターン）.xlsx"


def main() -> None:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE, OUT)

    wb = openpyxl.load_workbook(OUT)
    ws = wb[wb.sheetnames[0]]

    # ---- 表（B3:F10）を「3パターン + オプション」の一覧として利用 ----
    # 既存の書式（罫線・塗り・数値フォーマット）を壊さないため、値だけ差し替える
    rows = [
        # B(工程) / C(作業内容) / D(初期コスト) / E(3月末) / F(4月以降)
        ("パターン1（最低）", "アドバイザー（相談・レビュー中心）", 50000, "約5h/月", "追加稼働 10,000円/時間"),
        ("パターン2（中）", "アドバイザー + α（叩き台資料）", 100000, "約10h/月", "追加稼働 10,000円/時間"),
        ("パターン3（上）", "アドバイザー + α + β（図解/詳細化）", 200000, "約20h/月", "追加稼働 10,000円/時間"),
        ("オプション1", "ヒアリング強化（+5h）", 50000, "追加", ""),
        ("オプション2", "プレゼンテーション（+3h）", 30000, "追加", ""),
        ("オプション3", "フォローアップ（+5h）", 50000, "追加", ""),
    ]

    start_row = 4
    for i, (b, c, d, e, f) in enumerate(rows):
        r = start_row + i
        ws[f"B{r}"].value = b
        ws[f"C{r}"].value = c
        ws[f"D{r}"].value = d
        ws[f"E{r}"].value = e
        ws[f"F{r}"].value = f

    # 既存テンプレは行数6（D4:D9）を想定しているため、合計式を更新
    ws["D10"].value = "=SUM(D4:D9)"
    ws["E10"].value = ""
    ws["F10"].value = ""

    # ---- 月額費用（B12:F16）は今回は未使用（混乱を避けるためクリア）----
    for r in range(13, 16):
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{r}"].value = ""
    ws["D16"].value = ""

    # ---- 備考（B18）をMDの要旨で更新 ----
    note = "\n".join(
        [
            "【契約形態】",
            "・月額定額（リテイナー）／期間の定めなし（1ヶ月単位）",
            "・必要な月だけご利用いただく形式",
            "",
            "【換算単価（目安）】",
            "・1時間 10,000円",
            "",
            "【追加稼働】",
            "・10,000円/時間（必要に応じて、事前相談のうえ）",
            "",
            "【成果物】",
            "・月次ミーティング議事メモ（マークダウン）",
            "・パターン2以上：テーマ別整理資料（叩き台）※月1本目安",
            "・パターン3：叩き台 + 図解/詳細化（社内共有レベル）※月1〜2本目安",
            "",
            "【スケジュール目安】",
            "・2〜3月：相談・整理中心（負荷に合わせて稼働配分）",
            "・4月以降：テーマを進める（顧問先DB再設計、RPAクラウド化 等）",
        ]
    )
    ws["B18"].value = note

    wb.save(OUT)

    print(f"Written: {OUT}")


if __name__ == "__main__":
    main()

