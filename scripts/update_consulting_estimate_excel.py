# -*- coding: utf-8 -*-
"""対応範囲をビジネス寄りの表現に変更するスクリプト"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "税理士法人杉井総合会計事務所" / "提案" / "コンサルプラン見積もり（3パターン）.xlsx"

def update_summary_sheet(ws):
    """サマリーシートの対応範囲を更新"""
    # 対応範囲をよりビジネス寄りの表現に変更
    # 行9-25あたりが対応範囲
    
    # アドバイザー部分を簡潔に
    ws['A10'].value = "アドバイザー（相談・レビュー）"
    ws['A11'].value = "・月次の相談・進め方の整理"
    ws['A12'].value = "・設計・運用方法のレビュー・助言"
    ws['A13'].value = "・課題の整理・優先順位づけ"
    
    # 叩き台作成部分を簡潔に
    ws['A15'].value = "叩き台作成（簡易資料）"
    ws['A16'].value = "・課題・要件の整理資料作成"
    ws['A17'].value = "・改善方針の整理メモ"
    ws['A18'].value = "・優先順位・進め方のロードマップ"
    
    # 深掘り部分を簡潔に
    ws['A20'].value = "深掘り・図解・整備（詳細資料）"
    ws['A21'].value = "・現状整理（構成図・関係図）"
    ws['A22'].value = "・改善提案の詳細化・図解"
    ws['A23'].value = "・社内説明用資料の整備"

def update_pattern1_sheet(ws):
    """パターン1シートの対応範囲を更新"""
    ws['A1'].value = "パターン1: 最低（月5万）アドバイザープラン"
    
    # 概要説明を追加
    ws['A2'].value = "概要"
    ws['B2'].value = "迷いを減らす相談窓口としての支援プラン"
    
    ws['A3'].value = "対応範囲"
    
    # 対応範囲をビジネス寄りに
    ws['A4'].value = "アドバイザー（相談・レビュー中心）"
    ws['A5'].value = "・月次の相談・進め方の整理"
    ws['A6'].value = "・設計・運用方法のレビュー・助言"
    ws['A7'].value = "・課題の整理・優先順位づけ（口頭・メモレベル）"

def update_pattern2_sheet(ws):
    """パターン2シートの対応範囲を更新"""
    ws['A1'].value = "パターン2: 中（月10万）アドバイザー＋叩き台作成プラン"
    
    # 概要説明を追加
    ws['A2'].value = "概要"
    ws['B2'].value = "相談に加えて、叩き台を一緒に作って前に進めるプラン"
    
    ws['A3'].value = "対応範囲"
    
    # パターン1の全内容を含むことを明記
    ws['A4'].value = "パターン1の全内容"
    ws['A5'].value = ""
    ws['A6'].value = "+"
    ws['A7'].value = ""
    
    # 追加内容
    ws['A8'].value = "叩き台作成（簡易資料）"
    ws['A9'].value = "・課題・要件の整理資料作成（1テーマ）"
    ws['A10'].value = "・改善方針の整理メモ"
    ws['A11'].value = "・優先順位・進め方のロードマップ"

def update_pattern3_sheet(ws):
    """パターン3シートの対応範囲を更新"""
    ws['A1'].value = "パターン3: 上（月20万）アドバイザー＋叩き台作成＋詳細化プラン"
    
    # 概要説明を追加
    ws['A2'].value = "概要"
    ws['B2'].value = "相談 + 叩き台 + 図解/詳細化まで含め、社内合意形成に使える資料が残るプラン"
    
    ws['A3'].value = "対応範囲"
    
    # パターン2の全内容を含むことを明記
    ws['A4'].value = "パターン2の全内容"
    ws['A5'].value = ""
    ws['A6'].value = "+"
    ws['A7'].value = ""
    
    # 追加内容
    ws['A8'].value = "深掘り・図解・整備（詳細資料）"
    ws['A9'].value = "・現状整理（構成図・関係図）"
    ws['A10'].value = "・改善提案の詳細化・図解"
    ws['A11'].value = "・社内説明用資料の整備（実装前の判断材料）"

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # サマリーシートを更新
    if "サマリー" in wb.sheetnames:
        update_summary_sheet(wb["サマリー"])
    
    # 各パターンシートを更新
    if "パターン1" in wb.sheetnames:
        update_pattern1_sheet(wb["パターン1"])
    
    if "パターン2" in wb.sheetnames:
        update_pattern2_sheet(wb["パターン2"])
    
    if "パターン3" in wb.sheetnames:
        update_pattern3_sheet(wb["パターン3"])
    
    wb.save(EXCEL_PATH)
    print(f"Updated: {EXCEL_PATH}")

if __name__ == "__main__":
    main()
