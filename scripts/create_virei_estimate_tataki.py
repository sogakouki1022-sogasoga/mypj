# -*- coding: utf-8 -*-
"""
株式会社VIREI stAyle. 様向け 概算見積のたたき Excel を生成
・提案_請求 フォルダに 概算見積_たたき.xlsx を出力
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "株式会社VIREI stAyle.(ビレイスタイル)" / "提案_請求" / "概算見積_たたき.xlsx"

# 見積内訳: (フェーズ, 項目, 内容, 工数, 単価, 金額, 備考)
ROWS = [
    ("1.調査・要件整理", "連携調査", "LOYCUS×kintone連携可否・方式の調査・整理", 3, 50000, 150000, "API/仕様確認含む"),
    ("", "現状整理", "現状データ・フォーム・管理表の整理、要件のとりまとめ", 2, 50000, 100000, ""),
    ("", "小計", "", None, None, 250000, ""),
    ("", "", "", None, None, None, ""),
    ("2.kintone基本設計・構築", "基本設計", "顧客アプリを軸にしたアプリ構成・項目設計", 3, 50000, 150000, ""),
    ("", "アプリ構築", "顧客・問い合わせ・商品等のアプリ作成、関連レコード表示", 4, 50000, 200000, "標準機能中心"),
    ("", "小計", "", None, None, 350000, ""),
    ("", "", "", None, None, None, ""),
    ("3.LOYCUS連携", "連携設計・実装", "LOYCUS→kintoneデータ蓄積、ワンクリック連携等", 6, 50000, 300000, "※仕様確定後に精算"),
    ("", "（参考）", "※難易度により＋10～15万円の変動あり", None, None, None, ""),
    ("", "小計", "", None, None, 300000, ""),
    ("", "", "", None, None, None, ""),
    ("4.Shopify連携", "取込・突合", "Shopify顧客データのkintone取込、名寄せ・統一", 4, 50000, 200000, "方針確定後"),
    ("", "小計", "", None, None, 200000, ""),
    ("", "", "", None, None, None, ""),
    ("5.データ移行・運用", "データ移行", "Googleフォーム等の既存データ移行・マッピング", 2, 50000, 100000, "範囲により変動"),
    ("", "運用支援", "マニュアル・引き継ぎ、簡易トレーニング", 1, 50000, 50000, ""),
    ("", "小計", "", None, None, 150000, ""),
    ("", "", "", None, None, None, ""),
    ("", "", "合計（税別）", None, None, 1250000, ""),
    ("", "", "合計（税込10%）", None, None, 1375000, ""),
    ("", "", "", None, None, None, ""),
    ("【最小構成の目安】", "", "", None, None, None, ""),
    ("フェーズ1+2のみ（調査～kintone基本構築）", "顧客軸で情報を見る基盤のみ", "", None, None, 600000, "税別"),
    ("※LOYCUS・Shopify連携は別フェーズで追加", "", "", None, None, None, ""),
    ("", "", "", None, None, None, ""),
    ("【オプション】", "", "", None, None, None, ""),
    ("LOYCUS週次MTG同席（コンサル）", "都度 or 月額", "", None, None, "要相談", ""),
]


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "見積"

    thin = Side(style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 1行目: 宛先
    ws["A1"] = "（宛先）御中"
    ws["D1"] = "作成日："
    ws["E1"] = "有効期限：1ヶ月"
    ws.merge_cells("A1:C1")

    # 2行目 空
    # 3行目 タイトル
    ws["A3"] = "kintone導入支援 概算見積書"
    ws["A3"].font = Font(bold=True, size=12)
    ws.merge_cells("A3:G3")

    # 4-5 空
    # 6-7 前提条件
    ws["A6"] = "【前提条件】"
    ws["A6"].font = Font(bold=True)
    ws["A7"] = "・LOYCUS／Shopify連携の仕様は調査後に確定。本見積は想定範囲での概算です。"
    ws["A8"] = "・顧客情報の一元化（kintone集約）を軸に、段階導入を推奨します。"
    ws.merge_cells("A7:G7")
    ws.merge_cells("A8:G8")

    # 9 空
    # 10 表ヘッダー
    headers = ["フェーズ", "項目", "内容", "工数(人日)", "単価(円)", "金額(円)", "備考"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 11〜 データ
    for i, row in enumerate(ROWS, 11):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val if val != "" or c <= 3 else None)
            cell.border = border
            if c >= 4 and val is not None and val != "":
                cell.alignment = Alignment(horizontal="right" if c >= 5 else "left", vertical="top")
            if c in (2, 3):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 列幅
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18

    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
