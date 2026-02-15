# -*- coding: utf-8 -*-
"""
要件定義見積もり Excel 詳細版を生成（リトプラ docs に出力）
・各項目をサブ項目に分解し時間を振り分け
・成果物列を追加
・1シートで完結
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# リトプラ/docs に出力（mypj からの相対パス）
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(BASE, "リトプラ", "docs", "要件定義見積もり_詳細版.xlsx")
UNIT_PRICE = 10000

DETAILED_ROWS = [
    ("業務ヒアリング・現状分析", "業務フローヒアリング", "業務フローのヒアリング、現状課題の把握", "ヒアリング議事録", 10),
    ("業務ヒアリング・現状分析", "既存システム調査・分析", "既存システムの調査・分析、棚卸し前提の確認", "現状分析メモ", 8),
    ("業務ヒアリング・現状分析", "棚卸・店舗移動・在庫ロス等のヒアリング", "店舗移動・在庫ロス・破損・複数拠点運用のヒアリング", "ヒアリング議事録（追補）", 12),
    ("要件整理・機能要件定義", "機能要件の整理・機能一覧作成", "機能要件の整理、機能一覧の作成", "機能一覧表", 15),
    ("要件整理・機能要件定義", "優先順位決定・スマレジ連携範囲整理", "優先順位の決定、スマレジ連携範囲の整理", "優先度一覧・連携範囲表", 10),
    ("要件整理・機能要件定義", "商品/貯蔵品フロー整理", "商品・貯蔵品の流れの整理、要件の棚卸し", "業務フロー整理メモ", 10),
    ("データモデル設計", "アプリ構成・フィールド設計", "アプリ構成の検討、フィールド設計", "データモデル図・フィールド一覧", 18),
    ("データモデル設計", "データ連携設計", "kintone間・外部とのデータ連携設計", "連携設計メモ", 10),
    ("データモデル設計", "3PL/外部データ取込方式検討", "3PL/外部データのフォーマット検証・取込方式の検討", "取込仕様メモ", 7),
    ("業務フロー設計", "ToBe業務フロー図作成", "ToBe業務フロー図の作成、プロセス可視化", "ToBe業務フロー図", 15),
    ("業務フロー設計", "プロセス管理・入出荷履歴フロー", "プロセス管理設計、店舗移動・入出荷履歴のフロー定義", "プロセス定義書", 10),
    ("外部連携仕様定義", "freee・スマレジ・3PL連携調査", "freee・スマレジ・3PL等の連携可否・方式の調査", "連携調査結果表", 10),
    ("外部連携仕様定義", "API/CSV方式検討・連携項目確認", "API/CSV方式の検討、連携可能項目の確認", "連携仕様メモ", 10),
    ("デモ環境構築・画面検証", "要件検証用デモ環境整備", "要件検証用デモ環境の構築・整備", "デモ環境", 10),
    ("デモ環境構築・画面検証", "画面イメージ共有・レビュー対応", "画面イメージの共有、レビュー対応・指摘反映", "画面イメージ一覧・レビュー対応記録", 10),
    ("要件定義書・設計書作成", "要件定義書作成・レビュー対応", "要件定義書の作成、レビュー対応・指摘反映", "要件定義書", 18),
    ("要件定義書・設計書作成", "基本設計書作成・指摘反映", "基本設計書の作成、レビュー指摘の反映", "基本設計書", 17),
]


def style_header(ws):
    thin = Side(style="thin")
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def style_table(ws, max_row):
    thin = Side(style="thin")
    for r in range(1, max_row + 1):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if cell.border.left.style is None:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if c in (5, 6, 7) and r > 1 and cell.value is not None:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            if r > 1 and c in (2, 3, 4):
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "要件定義見積もり（詳細版）"

    headers = ["項目", "サブ項目", "内容", "成果物", "時間(h)", "単価(円)", "金額(税抜)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    last_data_row = len(DETAILED_ROWS) + 1  # データ最終行（例: 18）
    for i, (item, sub, desc, deliverable, hours) in enumerate(DETAILED_ROWS, 2):
        ws.cell(row=i, column=1, value=item)
        ws.cell(row=i, column=2, value=sub)
        ws.cell(row=i, column=3, value=desc)
        ws.cell(row=i, column=4, value=deliverable)
        ws.cell(row=i, column=5, value=hours)
        ws.cell(row=i, column=6, value=UNIT_PRICE)
        ws.cell(row=i, column=7, value=f"=E{i}*F{i}")  # 金額 = 時間×単価

    total_row = len(DETAILED_ROWS) + 2  # 合計行（例: 19）
    ws.cell(row=total_row, column=1, value="合計")
    for c in range(2, 5):
        ws.cell(row=total_row, column=c, value="")
    ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last_data_row})")   # 合計時間
    ws.cell(row=total_row, column=6, value=UNIT_PRICE)
    ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{last_data_row})")  # 合計金額
    for c in range(1, 8):
        ws.cell(row=total_row, column=c).font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14

    style_header(ws)
    style_table(ws, total_row)

    start_summary = total_row + 3
    summary = [
        ["要件定義フェーズ 見積サマリー", ""],
        ["", ""],
        ["合計時間", f'=E{total_row}&" 時間"'],   # 合計行を参照
        ["合計金額（税抜）", f"=G{total_row}"],   # 合計行を参照（セル書式で通貨表示）
        ["単価", f"¥{UNIT_PRICE:,}/h"],
        ["", ""],
        ["支払い条件", ""],
        ["要件定義完了時", "一括 100% お支払い"],
        ["（別案）", "着手金 30% / レビュー完了時 70%"],
        ["", ""],
        ["注意事項", ""],
        ["・上記金額は税抜表示です。", ""],
        ["・本見積もりはたたき台であり、要件定義完了後に開発フェーズの正式な見積もりを提示いたします。", ""],
    ]
    for r, row_data in enumerate(summary, start_summary):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # 合計金額（税抜）のセルは通貨書式
            if c == 2 and "=G" in str(val):
                cell.number_format = '"¥"#,##0'
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 55)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 28)
    for r in [start_summary, start_summary + 6, start_summary + 9]:
        ws.cell(row=r, column=1).font = Font(bold=True)

    wb.save(OUTPUT_FILE)
    print("Created:", OUTPUT_FILE)


if __name__ == "__main__":
    main()
